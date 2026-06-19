"""One-off converter: messy 'Fitness Tracker.xlsx' journal -> structured date,entry rows.

Rules (all documented so they're easy to change):
  - Dates become YYYY-MM-DD with a literal YYYY placeholder (year filled in later).
    The month carries forward until a new month is named.
  - Multi-set lines are split into one row per set group.
  - Bodyweight / assisted movements -> '@ 0kg' (assistance load is dropped).
  - 'bar' -> '@ 20kg' (standard Olympic bar; flagged in Review for you to confirm).
  - Cardio, prose notes and time-based holds (planks) go to the Review sheet, not
    the lifting rows.
  - Parenthetical comments are stripped from entries.
Anything that can't be parsed confidently is sent to Review rather than guessed.
"""
import csv
import datetime as dt
import re
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

SRC = "/Users/danielwelsh/Downloads/Fitness Tracker.xlsx"
OUT = "/Users/danielwelsh/Downloads/Fitness Tracker (formatted).xlsx"
OUT_CSV = "/Users/danielwelsh/Downloads/Fitness Tracker (import).csv"

BAR = 20.0  # assumed weight of an empty bar, in kg

MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june",
     "july", "august", "september", "october", "november", "december"], 1)}
ABBR = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12}
MONTH_LOOKUP = {**MONTHS, **ABBR}
MONTH_RE = "|".join(sorted(MONTH_LOOKUP, key=len, reverse=True))  # longest first
WEEKDAYS = "monday|tuesday|wednesday|thursday|friday|saturday|sunday"

CANON = {
    "squat": "Squat", "squats": "Squat",
    "bench": "Bench Press", "bench press": "Bench Press",
    "deadlift": "Deadlift", "deadlifts": "Deadlift",
    "ohp": "OHP", "overhead press": "OHP", "db ohp": "DB OHP", "dumbbell ohp": "DB OHP",
    "barbell row": "Barbell Row", "row": "Barbell Row", "rows": "Barbell Row", "bb row": "Barbell Row",
    "curl": "Curl", "curls": "Curl", "bicep curl": "Curl", "bicep curls": "Curl",
    "pull up": "Pull-ups", "pull ups": "Pull-ups", "pullup": "Pull-ups", "pullups": "Pull-ups",
    "pull-up": "Pull-ups", "pull-ups": "Pull-ups",
    "chin up": "Chin-ups", "chin ups": "Chin-ups", "chinup": "Chin-ups", "chinups": "Chin-ups",
    "chin-up": "Chin-ups", "chin-ups": "Chin-ups", "chin": "Chin-ups", "chins": "Chin-ups",
    "dip": "Dips", "dips": "Dips",
    "hanging leg raise": "Hanging Leg Raises", "hanging leg raises": "Hanging Leg Raises",
    "leg raise": "Hanging Leg Raises", "leg raises": "Hanging Leg Raises",
    "knee raise": "Knee Raises", "knee raises": "Knee Raises",
    "plank": "Plank",
}
BODYWEIGHT = {"Pull-ups", "Chin-ups", "Dips", "Hanging Leg Raises", "Knee Raises"}

_DATE_RE = re.compile(
    rf"^(?P<wd>{WEEKDAYS})?\s*(?P<day>\d{{1,2}})(?P<ord>st|nd|rd|th)?(?:\s+(?P<mon>{MONTH_RE}))?\b", re.I)
_DATE_RE_MF = re.compile(rf"^({MONTH_RE})\s+(\d{{1,2}})(?:st|nd|rd|th)?\b", re.I)


def parse_date(line):
    """Return (day, month|None) if the line *starts with* a date, else None.

    Requires a weekday, an ordinal suffix, or a month so a bare-number exercise
    line isn't mistaken for a date. Trailing text (e.g. "- 134lb") is allowed.
    """
    s = line.strip().lower()
    m = _DATE_RE.match(s)
    if m and (m.group("wd") or m.group("ord") or m.group("mon")):
        mon = MONTH_LOOKUP[m.group("mon")] if m.group("mon") else None
        return int(m.group("day")), mon
    m = _DATE_RE_MF.match(s)
    if m:
        return int(m.group(2)), MONTH_LOOKUP[m.group(1)]
    return None


def canon(name):
    key = re.sub(r"\s+", " ", name.strip().lower())
    if key in CANON:
        return CANON[key]
    n = name.strip()
    return n[:1].upper() + n[1:] if n else n


def fmtw(w):
    return str(int(w)) if w == int(w) else ("%.2f" % w).rstrip("0").rstrip(".")


def to_entries(line):
    """Return (entries:list[str], reason:str|None). reason set => goes to Review."""
    low = line.lower()
    if re.search(r"\b(running machine|treadmill|cycle|cycling|bike|jog|swim|elliptical|cardio)\b", low):
        return [], "cardio (use Garmin)"
    if re.search(r"\brun\b|\bmph\b|mpkm|per km|/km|\bmile", low) and not re.search(r"\brow", low):
        return [], "cardio (use Garmin)"
    if re.search(r"\bplank\b|\b(sec|secs|second|minute|hold)\b", low):
        return [], "time-based (handle manually)"

    s = re.sub(r"\([^)]*\)", "", line).strip().rstrip(".").strip()

    # Exercise name = leading letters before the first number / @ / ~.
    nm = re.match(r"^([A-Za-z][A-Za-z \-/&.]*?)\s*(?=[\d@~])", s)
    if not nm:
        return [], "note / prose (skipped)"
    name = nm.group(1).strip(" ,")
    rest = s[nm.end():]

    bar_in_name = bool(re.search(r"\bbar\b", name, re.I))
    if bar_in_name:
        name = re.sub(r"\bbar\b", "", name, flags=re.I).strip()
    cname = canon(name)
    if not cname:
        return [], "no exercise name"

    # Approx reps ("~12 reps") -> a single set group.
    rest = re.sub(r"~\s*(\d+)\s*reps?", r"1x\1", rest, flags=re.I)

    groups = [(int(a), int(b)) for a, b in re.findall(r"(\d+)\s*x\s*(\d+)", rest)]
    if not groups:
        return [], "no set groups parsed"

    # Weights: prefer explicit '@ weight' / 'bar'; else a bare leading number
    # (this journal often writes "Bench 37.5 2x5 1x7" with no @).
    weights = [BAR if w.group(1).lower() == "bar" else float(w.group(1))
               for w in re.finditer(r"@\s*(bar|\d+(?:\.\d+)?)", rest, re.I)]
    if re.search(r"@\s*bar|\bbar\b", rest, re.I) and not weights:
        weights = [BAR]
    if not weights:
        lead = re.match(r"^\s*(\d+(?:\.\d+)?)(?!\s*x)\s*(?:kg)?\b", rest, re.I)
        if lead:
            weights = [float(lead.group(1))]
    if not weights:
        # weight written after the sets, e.g. "Bench 3x5 45"
        leftover = re.sub(r"\d+\s*x\s*\d+", " ", rest)
        tw = re.fullmatch(r"[\s,]*(\d+(?:\.\d+)?)\s*(?:kg)?[\s,]*", leftover)
        if tw:
            weights = [float(tw.group(1))]
    if bar_in_name and not weights:
        weights = [BAR]
    if not weights and re.search(r"unweighted|bodyweight|body weight|\bbw\b", low):
        weights = [0.0]

    if cname in BODYWEIGHT:
        assigned = [0.0] * len(groups)
    elif not weights:
        return [], "no weight found (and not a known bodyweight move)"
    elif len(weights) == len(groups):
        assigned = weights                      # e.g. "3x5 @ bar, 1x3 @ 22.5"
    else:
        assigned = [weights[0]] * len(groups)   # one weight applies to all groups

    entries = [f"{cname} {st}x{rp} @ {fmtw(w)}kg" for (st, rp), w in zip(groups, assigned)]
    return entries, None


def real_year(pre, mm):
    """Map a Y# block to the real calendar year the user confirmed.

    Y3 spans a long gap, so it splits: Jan–Mar is 2024, Oct–Dec is 2025.
    """
    return {
        "Y1": 2022,
        "Y2": 2023,
        "Y3": 2024 if mm <= 3 else 2025,
        "Y4": 2026,
    }.get(pre)


def with_real_year(datestr):
    if not datestr or not datestr.startswith("Y"):
        return datestr
    pre, mm, dd = datestr.split("-")
    y = real_year(pre, int(mm))
    return f"{y}-{mm}-{dd}" if y else datestr


def main():
    ws = load_workbook(SRC, data_only=True)["Sheet1"]
    values = [c[0].value for c in ws.iter_rows()]

    workouts, review = [], []
    cur_day = cur_mon = None
    cur_date = ""
    seg, prev_mon, rollovers = 1, None, []
    stats = {"entries": 0, "review": 0, "dates": 0}

    for v in values:
        if v is None:
            continue

        day = mon = None
        if isinstance(v, (dt.datetime, dt.date)):
            # Real Excel date cell. Use month/day only — the cell years are
            # unreliable (many are future-dated), so it joins the Yn scheme too.
            real = v.date() if isinstance(v, dt.datetime) else v
            day, mon = real.day, real.month
        elif isinstance(v, (int, float)):
            if 1990 <= v <= 2100:        # bare year number = section marker
                continue
            v = str(v)

        if day is None:                  # string cell: week header, text date, or content
            raw = str(v).strip()
            if not raw:
                continue
            if re.match(r"(?i)^week\s+\d+$", raw):
                continue
            pd = parse_date(raw)
            if pd:
                day, mon = pd
            else:
                entries, reason = to_entries(raw)
                if reason and not entries:
                    review.append((cur_date, raw, reason))
                    stats["review"] += 1
                else:
                    for e in entries:
                        workouts.append((cur_date, e, raw))
                        stats["entries"] += 1
                continue

        # We have a date header (from a real date cell or a parsed text date).
        if mon:
            cur_mon = mon
        cur_day = day
        if cur_mon is not None:
            if prev_mon is not None and cur_mon < prev_mon:   # month backwards => new year block
                seg += 1
                rollovers.append(f"Y{seg} begins around month {cur_mon:02d}, day {cur_day}")
            prev_mon = cur_mon
        cur_date = f"Y{seg}-{cur_mon:02d}-{cur_day:02d}" if cur_mon and cur_day else ""
        stats["dates"] += 1

    # Apply the confirmed real years (Y3 splits across the 2024/2025 gap).
    workouts = [(with_real_year(d), e, s) for d, e, s in workouts]
    review = [(with_real_year(d), e, s) for d, e, s in review]

    seg_md = {}
    for d, _, _ in workouts:
        pre, md = d.split("-", 1)
        seg_md.setdefault(pre, []).append(md)
    seg_lines = [f"{k}: {min(v)} .. {max(v)}" for k, v in sorted(seg_md.items())]

    # Session timeline with gaps, to reveal long breaks (and possible hidden years).
    per_date = Counter(d for d, _, _ in workouts)
    order, seen = [], set()
    for d, _, _ in workouts:
        if d not in seen:
            seen.add(d)
            order.append(d)
    timeline, prev = [], None
    for d in order:
        pre, mm, dd = d.split("-")
        doy = dt.date(2001, int(mm), int(dd)).timetuple().tm_yday
        if prev and prev[0] == pre:
            gap = doy - prev[1]
            note = "out-of-order (check year)" if gap < 0 else ("long gap" if gap >= 28 else "")
            timeline.append((d, per_date[d], gap, note))
        else:
            timeline.append((d, per_date[d], "", "new year block" if prev else "start"))
        prev = (pre, doy)

    wb = Workbook()
    sh = wb.active
    sh.title = "Workouts"
    sh.append(["date", "entry", "source_line"])
    for r in workouts:
        sh.append(list(r))
    rv = wb.create_sheet("Review")
    rv.append(["date", "original_line", "reason"])
    for r in review:
        rv.append(list(r))

    readme = wb.create_sheet("ReadMe", 0)
    notes = [
        "Fitness Tracker — converted to date, entry format",
        "",
        "HOW TO USE",
        "1. Dates are real years (YYYY-MM-DD), confirmed with you. Ranges in this file:",
    ] + ["      " + s for s in seg_lines] + [
        "   Note: there was a long training gap from Mar 2024 to Oct 2025.",
        "   The 'Timeline' tab shows the gaps between sessions.",
        "2. Copy the 'date' and 'entry' columns (A and B) from 'Workouts' into your",
        "   Google Sheet. The scraper reads date + entry; 'source_line' is just for",
        "   your reference and can be ignored.",
        "3. 'Review' lists lines I could not convert automatically (notes, cardio,",
        "   incomplete sets) — add any you want by hand.",
        "",
        "CONVERSION RULES",
        "- Multi-set lines were split into one row each.",
        "- Bodyweight / assisted moves -> '@ 0kg'.",
        "- 'bar' -> '@ 20kg' (assumed Olympic bar; adjust if yours differs).",
        "- Cardio, planks/time-based, and prose notes went to 'Review'.",
        "- All weights are kilograms.",
    ]
    for line in notes:
        readme.append([line])

    tl = wb.create_sheet("Timeline")
    tl.append(["date", "exercises_logged", "gap_days", "note"])
    for row in timeline:
        tl.append(list(row))

    for sheet in (sh, rv, tl):
        for cell in sheet[1]:
            cell.font = Font(name="Arial", bold=True)
    for col, width in (("A", 14), ("B", 46), ("C", 46)):
        sh.column_dimensions[col].width = width
        rv.column_dimensions[col].width = width
    for col, width in (("A", 14), ("B", 16), ("C", 10), ("D", 16)):
        tl.column_dimensions[col].width = width
    readme.column_dimensions["A"].width = 78
    for cell in readme["A"]:
        cell.font = Font(name="Arial")
    wb.save(OUT)

    # Clean date,entry CSV ready to import into the Google Sheet.
    with open(OUT_CSV, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date", "entry"])
        writer.writerows((d, e) for d, e, _ in workouts)
    print(f"Wrote {OUT_CSV}")

    longest = sorted((t for t in timeline if isinstance(t[2], int)), key=lambda t: -t[2])[:6]
    print("\nLongest within-block gaps:")
    for d, n, g, note in longest:
        print(f"  {g:>3} days before {d}")
    print("New-year block starts:", [t[0] for t in timeline if t[3] == "new year block"])
    ooo = [t[0] for t in timeline if t[3] == "out-of-order (check year)"]
    if ooo:
        print("Out-of-order dates (likely year typos in source):", ooo)

    print(f"Wrote {OUT}")
    print(f"  workout rows: {stats['entries']}   review rows: {stats['review']}   dates seen: {stats['dates']}")
    print(f"  year segments detected: {seg}")
    for r in rollovers:
        print(f"    - {r}")
    print("\n--- first 25 Workouts ---")
    for r in workouts[:25]:
        print(f"  {r[0]:<12} {r[1]:<34} | {r[2]}")
    print("\n--- Review reason counts + samples ---")
    by_reason = {}
    for r in review:
        by_reason.setdefault(r[2], []).append(r[1])
    for reason, n in Counter(r[2] for r in review).most_common():
        print(f"\n  [{n}] {reason}")
        for sample in by_reason[reason][:6]:
            print(f"       {sample}")


if __name__ == "__main__":
    main()
